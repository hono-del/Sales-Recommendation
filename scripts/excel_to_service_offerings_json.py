"""
記入済み Excel を config/service-offerings.json に変換する（集約担当者向け）。

使い方:
  py scripts/excel_to_service_offerings_json.py data/templates/service-offerings-catalog.xlsx
  py scripts/excel_to_service_offerings_json.py path/to/山田_サービス案.xlsx --merge
"""
from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "config" / "service-offerings.json"
HEADER_ROW = 1
LABEL_ROW = 2
DATA_START_ROW = 4  # 3行目は記入例


def _row_to_offering(row: list, headers: list[str]) -> dict | None:
    data = {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}
    raw_id = data.get("id")
    if not raw_id or str(raw_id).strip() == "":
        return None
    if str(raw_id).startswith("svc_example"):
        return None

    def _need_list(prefix: str, weight: float) -> list[dict]:
        out = []
        for key in (f"{prefix}_1", f"{prefix}_2", f"{prefix}_3"):
            name = data.get(key)
            if name and str(name).strip():
                out.append({"name": str(name).strip(), "weight": weight})
        return out

    primary = _need_list("primary_need", 1.0)
    if not primary:
        raise ValueError(f"id={raw_id}: primary_need_1 は必須です")

    secondary = []
    for key in ("secondary_need_1", "secondary_need_2"):
        name = data.get(key)
        if name and str(name).strip():
            secondary.append({"name": str(name).strip(), "weight": 0.5})

    load_labels = []
    for key in ("load_label_1", "load_label_2"):
        v = data.get(key)
        if v and str(v).strip():
            load_labels.append(str(v).strip())

    value_axes = []
    for key in ("value_axis_1", "value_axis_2"):
        v = data.get(key)
        if v and str(v).strip():
            value_axes.append(str(v).strip())

    burden = []
    for key in ("burden_addressed_1", "burden_addressed_2"):
        v = data.get(key)
        if v and str(v).strip():
            burden.append(str(v).strip())

    value_shift = []
    for key in ("value_shift_1", "value_shift_2"):
        v = data.get(key)
        if v and str(v).strip():
            value_shift.append(str(v).strip())

    analog_name = data.get("analog_name")
    analog_pattern = data.get("analog_pattern")
    analog = None
    if (analog_name and str(analog_name).strip()) or (analog_pattern and str(analog_pattern).strip()):
        analog = {
            "name": str(analog_name or "（社内オリジナル）").strip(),
            "pattern": str(analog_pattern or "other").strip(),
            "why": str(data.get("analog_why") or "").strip(),
            "pitch_template": str(data.get("pitch_template") or "").strip(),
        }

    return {
        "id": str(raw_id).strip(),
        "title": str(data.get("title") or "").strip(),
        "one_liner": str(data.get("one_liner") or "").strip(),
        "direction": str(data.get("direction") or "").strip(),
        "domain": str(data.get("domain") or "").strip(),
        "lifecycle": str(data.get("lifecycle") or "").strip(),
        "status": str(data.get("status") or "draft").strip(),
        "contributor": str(data.get("contributor") or "").strip(),
        "updated_at": str(data.get("updated_at") or date.today().isoformat()).strip(),
        "primary_needs": primary,
        "secondary_needs": secondary,
        "load_labels": load_labels,
        "value_axes": value_axes,
        "burden_addressed": burden,
        "value_shift": value_shift,
        "need_rationale": str(data.get("need_rationale") or "").strip(),
        "trade_off": str(data.get("trade_off") or "").strip(),
        "analog": analog,
        "eligibility": str(data.get("eligibility") or "").strip() or None,
        "combinability": str(data.get("combinability") or "").strip() or None,
        "notes": str(data.get("notes") or "").strip(),
    }


def load_offerings_from_excel(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    if "記入" not in wb.sheetnames:
        raise SystemExit(f"シート「記入」がありません: {path}")
    ws = wb["記入"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < DATA_START_ROW:
        return []
    headers = [str(c).strip() if c else "" for c in rows[HEADER_ROW - 1]]
    offerings = []
    for row in rows[DATA_START_ROW - 1 :]:
        if not any(row):
            continue
        off = _row_to_offering(list(row), headers)
        if off:
            offerings.append(off)
    return offerings


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("excel", type=Path, help="入力 Excel（記入シート）")
    parser.add_argument(
        "--merge",
        action="store_true",
        help="既存 service-offerings.json に追記・上書き（同 id は上書き）",
    )
    parser.add_argument("-o", "--output", type=Path, default=OUT_PATH)
    args = parser.parse_args()

    new_items = load_offerings_from_excel(args.excel)
    if args.merge and args.output.exists():
        existing = json.loads(args.output.read_text(encoding="utf-8"))
        by_id = {o["id"]: o for o in existing.get("offerings", [])}
        for o in new_items:
            by_id[o["id"]] = o
        payload = {"version": existing.get("version", "1.0"), "description": existing.get("description", ""), "offerings": list(by_id.values())}
    else:
        payload = {
            "version": "1.0",
            "description": "アップグレード/ダウングレードサービス案カタログ（正本）",
            "offerings": new_items,
        }

    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(payload['offerings'])} offerings -> {args.output}")


if __name__ == "__main__":
    main()
