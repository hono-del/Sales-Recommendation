"""
社内サービス案カタログ用 Excel テンプレートを生成する。
実行: py scripts/build_service_offerings_excel.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = ROOT / "data" / "templates" / "service-offerings-catalog.xlsx"

# Need マスタ（graph_builder.py NEW_NEED_MASTER と同期）
NEEDS = [
    ("ChildSafety", "子供を安全に乗せたい"),
    ("EasyChildPickup", "子供の乗せ降ろしを楽にしたい"),
    ("FamilyConversation", "家族で会話しやすい空間が欲しい"),
    ("FamilyComfort", "家族全員が快適に移動したい"),
    ("LargeCargoForFamily", "ベビーカーや荷物を楽に積みたい"),
    ("WeekendFamilyTrip", "家族旅行を楽しみたい"),
    ("ChildMonitoringEase", "後席の子供の様子を確認したい"),
    ("StressFreeSchoolPickup", "送迎のストレスを減らしたい"),
    ("PetFriendlyTravel", "ペットと快適に移動したい"),
    ("EasyParking", "駐車を楽にしたい"),
    ("VisibilityConfidence", "周囲を見やすくしたい"),
    ("DrivingConfidence", "運転への不安を減らしたい"),
    ("AccidentAnxietyReduction", "事故不安を減らしたい"),
    ("NightDrivingConfidence", "夜間運転を安心したい"),
    ("SnowDrivingConfidence", "雪道でも安心したい"),
    ("BeginnerFriendlyDriving", "運転初心者でも扱いやすくしたい"),
    ("ElderlyDrivingSupport", "高齢でも安全に運転したい"),
    ("FatigueReduction", "長時間運転で疲れたくない"),
    ("QuietCabinExperience", "静かな空間で移動したい"),
    ("SmoothRideComfort", "乗り心地を良くしたい"),
    ("StressFreeCommute", "通勤ストレスを減らしたい"),
    ("RelaxingDrive", "リラックスして運転したい"),
    ("ComfortableLongDistanceTravel", "長距離移動を快適にしたい"),
    ("ClimateComfort", "車内温度を快適に保ちたい"),
    ("FlexibleCargoSpace", "荷物量に柔軟対応したい"),
    ("OutdoorGearTransport", "アウトドア用品を積みたい"),
    ("EasyLoading", "荷物を積み下ろししやすくしたい"),
    ("LargeShoppingCapacity", "まとめ買いに対応したい"),
    ("SportsEquipmentTransport", "スポーツ用品を運びたい"),
    ("FlatSeatUtility", "車中泊や大きな荷物に対応したい"),
    ("LowFuelAnxiety", "燃料代不安を減らしたい"),
    ("MaintenanceCostReduction", "維持費を抑えたい"),
    ("LongTermReliability", "長く安心して乗りたい"),
    ("ResaleValueRetention", "リセール価値を維持したい"),
    ("EfficientDailyMobility", "日常移動コストを抑えたい"),
    ("DrivingEnjoyment", "運転そのものを楽しみたい"),
    ("AdventureLifestyle", "冒険感を楽しみたい"),
    ("OutdoorLifestyle", "アウトドア生活を楽しみたい"),
    ("EmotionalAttachment", "愛着を持てる車に乗りたい"),
    ("PersonalExpression", "自分らしさを表現したい"),
    ("PremiumFeeling", "上質感を感じたい"),
    ("ExcitingAcceleration", "加速感を楽しみたい"),
    ("StatusRecognition", "周囲から良く見られたい"),
    ("UrbanManeuverability", "狭い道でも扱いやすくしたい"),
    ("CompactParkingEase", "小さい駐車場でも停めやすくしたい"),
    ("ShortTripEfficiency", "短距離移動を効率化したい"),
    ("QuickErrandMobility", "ちょっとした移動を楽にしたい"),
    ("EasyEntryExit", "乗り降りを楽にしたい"),
    ("LowPhysicalBurden", "身体負担を減らしたい"),
    ("AccessibleSeating", "座りやすい姿勢を確保したい"),
    ("CaregiverSupport", "介護・送迎を楽にしたい"),
    ("EnvironmentalResponsibility", "環境配慮したい"),
    ("ChargingConfidence", "充電不安を減らしたい"),
    ("QuietElectricExperience", "EVの静かさを楽しみたい"),
    ("EnergyEfficiency", "エネルギー効率を高めたい"),
]

LOAD_LABELS = [
    "長距離移動による疲労",
    "渋滞ストレス",
    "駐車・狭い道への不安",
    "操作の難しさ",
    "情報過多による判断負荷",
    "家族同乗時の不満リスク",
    "維持費への不安",
    "機能不足による後悔",
    "使わない設備への投資",
    "すぐ飽きるリスク",
]

COLUMNS = [
    ("id", "案ID（英小文字・数字・_）", True),
    ("title", "サービス名", True),
    ("one_liner", "一言（20〜40字）", True),
    ("direction", "方向 upgrade/downgrade/neutral", True),
    ("domain", "領域", True),
    ("lifecycle", "ライフサイクル", True),
    ("status", "ステータス", False),
    ("contributor", "記入者", True),
    ("updated_at", "更新日 YYYY-MM-DD", False),
    ("primary_need_1", "主Need①（英語name）", True),
    ("primary_need_2", "主Need②", False),
    ("primary_need_3", "主Need③", False),
    ("secondary_need_1", "副Need①", False),
    ("secondary_need_2", "副Need②", False),
    ("load_label_1", "負荷ラベル①", False),
    ("load_label_2", "負荷ラベル②", False),
    ("value_axis_1", "価値観軸①", False),
    ("value_axis_2", "価値観軸②", False),
    ("burden_addressed_1", "軽減する負荷①", False),
    ("burden_addressed_2", "軽減する負荷②", False),
    ("value_shift_1", "価値の変化①", False),
    ("value_shift_2", "価値の変化②", False),
    ("need_rationale", "Need選定理由", True),
    ("trade_off", "トレードオフ（1文）", True),
    ("analog_name", "模倣する実在サービス名", False),
    ("analog_pattern", "模倣パターン", False),
    ("analog_why", "模倣理由", False),
    ("pitch_template", "ユーザー向け例文", False),
    ("eligibility", "対象者", False),
    ("combinability", "併用 stackable/exclusive", False),
    ("notes", "メモ", False),
]

CHOICES = {
    "direction": ["upgrade", "downgrade", "neutral"],
    "domain": [
        "maintenance",
        "connectivity",
        "insurance_warranty",
        "ownership_program",
        "upgrade_path",
        "downgrade_path",
        "trade_lifecycle",
        "concierge_support",
        "other",
    ],
    "lifecycle": ["pre_purchase", "ownership", "renewal", "exit"],
    "status": ["draft", "reviewed", "approved"],
    "value_axis": ["safety", "family", "efficiency", "enjoyment", "adventure"],
    "value_shift": [
        "cost_down",
        "risk_down",
        "convenience_up",
        "prestige_up",
        "control_up",
        "simplicity_up",
    ],
    "analog_pattern": [
        "predictable_cost",
        "good_enough_bundle",
        "tier_change",
        "expert_curated",
        "self_serve_compare",
        "usage_based",
        "insurance_wrapper",
        "loyalty_retention",
        "other",
    ],
    "eligibility": ["新規購入者", "既存オーナー", "契約満了時", "制限なし"],
    "combinability": ["stackable", "exclusive"],
}

EXAMPLE_ROW = {
    "id": "svc_example_fixed_maintenance",
    "title": "【記入例】定額メンテナンスプラン",
    "one_liner": "月額固定で、突然の整備費不安を減らす",
    "direction": "upgrade",
    "domain": "maintenance",
    "lifecycle": "ownership",
    "status": "draft",
    "contributor": "山田太郎",
    "updated_at": "2026-05-28",
    "primary_need_1": "MaintenanceCostReduction",
    "primary_need_2": "LongTermReliability",
    "primary_need_3": "",
    "secondary_need_1": "",
    "secondary_need_2": "",
    "load_label_1": "維持費への不安",
    "load_label_2": "",
    "value_axis_1": "efficiency",
    "value_axis_2": "safety",
    "burden_addressed_1": "維持費への不安",
    "burden_addressed_2": "",
    "value_shift_1": "risk_down",
    "value_shift_2": "simplicity_up",
    "need_rationale": "維持費の見通しが立つことで、長く乗る安心感にもつながる",
    "trade_off": "月額は増えるが、突発的な整備費の不安が減る",
    "analog_name": "携帯電話の定額プラン",
    "analog_pattern": "predictable_cost",
    "analog_why": "変動費を固定費化して判断負荷を下げる",
    "pitch_template": "携帯の定額プランのように、クルマの維持費を毎月一定に",
    "eligibility": "既存オーナー",
    "combinability": "stackable",
    "notes": "記入後この行は削除可",
}


def _header_style():
    return Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")


def _build_choices_sheet(wb: Workbook) -> dict[str, str]:
    ws = wb.create_sheet("選択肢", 1)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ranges: dict[str, str] = {}

    row = 1
    for key, values in CHOICES.items():
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        start = row + 1
        row += 1
        for v in values:
            ws.cell(row=row, column=1, value=v)
            row += 1
        ranges[key] = f"=選択肢!$A${start}:$A${row - 1}"
        row += 1

    ws.cell(row=row, column=1, value="need_name").font = Font(bold=True)
    ws.cell(row=row, column=2, value="need_label").font = Font(bold=True)
    row += 1
    need_start = row
    for name, label in NEEDS:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=label)
        row += 1
    ranges["need"] = f"=選択肢!$A${need_start}:$A${row - 1}"

    row += 1
    ws.cell(row=row, column=1, value="load_label").font = Font(bold=True)
    row += 1
    load_start = row
    for label in LOAD_LABELS:
        ws.cell(row=row, column=1, value=label)
        row += 1
    ranges["load"] = f"=選択肢!$A${load_start}:$A${row - 1}"

    wb._validation_ranges = ranges  # type: ignore[attr-defined]
    return ranges


def _add_list_validation(ws, col_letter: str, formula: str, start_row: int = 3, end_row: int = 502) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "リストから選んでください"
    dv.errorTitle = "入力エラー"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def _build_entry_sheet(wb: Workbook, ranges: dict[str, str]) -> None:
    ws = wb.active
    ws.title = "記入"

    font_bold, fill = _header_style()
    for col_idx, (field, label, required) in enumerate(COLUMNS, start=1):
        mark = "※" if required else ""
        c1 = ws.cell(row=1, column=col_idx, value=field)
        c2 = ws.cell(row=2, column=col_idx, value=f"{label}{mark}")
        for c in (c1, c2):
            c.font = font_bold
            c.fill = fill
            c.alignment = Alignment(wrap_text=True, vertical="center")
        width = max(12, min(36, len(label) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"
    field_to_col = {field: get_column_letter(i) for i, (field, _, _) in enumerate(COLUMNS, start=1)}

    for col_idx, (field, _, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=3, column=col_idx, value=EXAMPLE_ROW.get(field, ""))

    field_range_map = {
        "direction": "direction",
        "domain": "domain",
        "lifecycle": "lifecycle",
        "status": "status",
        "value_axis_1": "value_axis",
        "value_axis_2": "value_axis",
        "value_shift_1": "value_shift",
        "value_shift_2": "value_shift",
        "analog_pattern": "analog_pattern",
        "eligibility": "eligibility",
        "combinability": "combinability",
        "primary_need_1": "need",
        "primary_need_2": "need",
        "primary_need_3": "need",
        "secondary_need_1": "need",
        "secondary_need_2": "need",
        "load_label_1": "load",
        "load_label_2": "load",
        "burden_addressed_1": "load",
        "burden_addressed_2": "load",
    }
    for field, range_key in field_range_map.items():
        _add_list_validation(ws, field_to_col[field], ranges[range_key])


def _build_help_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("使い方", 2)
    lines = [
        "【社内サービス案カタログ】Excel 記入シート",
        "",
        "1. 「記入」シートの4行目以降に1案1行で入力（3行目は記入例）",
        "2. プルダウンがある列はリストから選択（Need・負荷ラベル等）",
        "3. ※付き列は必須（2行目のラベル参照）",
        "4. 記入例行（3行目）は完了後削除してOK",
        "5. 提出: このファイルをそのまま共有（ファイル名に担当者名を入れるとよい）",
        "",
        "詳細手順・判断基準: docs/SERVICE_OFFERING_CATALOG_WORK_GUIDE.md",
        "",
        "列の意味（英語フィールド名は1行目）",
        "・direction: upgrade=体験向上系 / downgrade=コスト最適化系 / neutral=横移動",
        "・primary_need: 生活欲求（英語name）。主1〜3件",
        "・load_label: デモの負荷ラベルと同じ文言を選択",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 80


def main() -> None:
    wb = Workbook()
    ranges = _build_choices_sheet(wb)
    _build_entry_sheet(wb, ranges)
    _build_help_sheet(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
